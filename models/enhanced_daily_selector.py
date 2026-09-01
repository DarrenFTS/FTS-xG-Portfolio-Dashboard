"""
FTS xG Portfolio — Enhanced Daily Selector Excel Export
Formatting matches FTS_Selections reference file:

  Main "Selections" tab — 4 LIVE systems only (Lay U1.5, Back O2.5, Lay O3.5, FHG Lay U0.5)
  "Test" tab — Back the Draw only, same layout/formulas, separate sheet

  B-F  : Calibri, size 11, black font (000000), no fill override — plain data columns
  G    : system dark bg, white, bold=True, size 11, center           (Market)
  H    : system light bg, system fg, bold=True, number_format 0.00   (6G xG)
  I    : system light bg, system fg, bold=False, size 11, center     (Rule)
  J    : alternating F2F6FB/FFFFFF, fg=1A5C9E, bold=True, number_format 0.00  (Odds)
  K    : alternating F2F6FB/FFFFFF, fg=0B5E6B, bold=True, size 11    (Hist ROI)
  L-O  : system columns (4 live systems), number_format 0.00, inactive=F0F0F0/CCCCCC
  P    : Row Total, number_format 0.00
  Q    : Month, number_format 0.00
  R    : Cumulative, number_format 0.00
  S    : £, number_format 0.00
"""
import os, sys
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from systems.all_systems import BetSignal

LIVE_SYS = {
    "Lay U1.5":      {"col": 12, "light": "D4EEF2", "dark": "0B5E6B", "fg": "0B5E6B"},
    "Back O2.5":     {"col": 13, "light": "D6EFE1", "dark": "217346", "fg": "217346"},
    "Lay O3.5":      {"col": 14, "light": "EBE0F0", "dark": "4A235A", "fg": "4A235A"},
    "FHG Lay U0.5":  {"col": 15, "light": "FFF0DC", "dark": "B35C00", "fg": "B35C00"},
}
TEST_SYS = {
    "Back the Draw": {"col": 12, "light": "D6EAF8", "dark": "1A5C9E", "fg": "1A5276"},
}

NAVY       = "0D2B55"
INACT_BG   = "F0F0F0"
INACT_FG   = "CCCCCC"
ACTIVE_FG  = "0D2B55"
JK_ODD_BG  = "F2F6FB"
JK_EVEN_BG = "FFFFFF"
J_FG       = "1A5C9E"
K_FG_POS   = "0B5E6B"
BODY_FONT  = "Calibri"
BODY_BLACK = "000000"

def _bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _plain(ws, row, col, val, size, align, number_format=None):
    """B-F style: Calibri, size 11, black font, no fill override, plain data cell."""
    c = ws.cell(row, col, val if val is not None else "")
    c.font      = Font(name=BODY_FONT, bold=False, size=size, color=BODY_BLACK)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _bdr()
    if number_format:
        c.number_format = number_format
    return c

def _c(ws, row, col, val, bg, fg, bold, size, align, number_format=None):
    c = ws.cell(row, col, val if val is not None else "")
    c.font      = Font(name=BODY_FONT, bold=bold, size=size, color=fg)
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _bdr()
    if number_format:
        c.number_format = number_format
    return c

def _fmt_date(val):
    if not val or str(val).strip() in ('', 'nan', 'None'):
        return ''
    try:
        if isinstance(val, datetime):
            return val
        s = str(val).strip().split(' ')[0].split('T')[0]
        return datetime.strptime(s, '%Y-%m-%d') if '-' in s else s
    except Exception:
        return str(val)

def _fmt_time(val):
    if not val or str(val).strip() in ('', 'nan', 'None'):
        return ''
    s = str(val).strip()
    if ' ' in s:
        s = s.split(' ')[-1]
    parts = s.split(':')
    return f"{parts[0]}:{parts[1]}" if len(parts) >= 2 else s


def _write_sheet(ws, signals, date_str, sys_map, sheet_title):
    """Write one sheet (Selections or Test) with the given signals and system map."""
    ws.sheet_view.showGridLines = False
    n_sys = len(sys_map)
    first_sys_col = 12
    last_sys_col  = first_sys_col + n_sys - 1
    row_total_col = last_sys_col + 1
    month_col     = row_total_col + 1
    cum_col       = month_col + 1
    gbp_col       = cum_col + 1

    col_widths = {'A':2,'B':12,'C':8,'D':25,'E':20,'F':13,'G':17,'H':8,'I':21.5,'J':8,'K':10}
    for i in range(n_sys):
        col_widths[chr(ord('L') + i)] = 14.8 if i == 0 else 13
    extra_cols = ['Q','R','S','T'][:4]
    for i, letter in enumerate(['Row Total','Month','Cumulative','£']):
        pass
    widths_tail = [10, 13, 12, 8.7]
    for i, w in enumerate(widths_tail):
        col_letter = chr(ord('A') + row_total_col - 1)
        col_widths[col_letter] = w
        row_total_col_letter = None

    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
    # Set widths for the tail columns explicitly by index
    from openpyxl.utils import get_column_letter
    for offset, w in zip(range(0, 4), widths_tail):
        ws.column_dimensions[get_column_letter(row_total_col + offset)].width = w

    # ── Row 1: Title + system total formulas ──────────────────────────────────
    ws.row_dimensions[1].height = 21.75
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=11)
    c = ws.cell(1, 2)
    c.value = f"{sheet_title}  —  {date_str}"
    c.font  = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill  = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")

    from openpyxl.utils import get_column_letter
    for i, (sys_name, cfg) in enumerate(sys_map.items()):
        col_num = first_sys_col + i
        col_letter = get_column_letter(col_num)
        formula = f"=SUM({col_letter}3:{col_letter}200)"
        c = ws.cell(1, col_num, formula)
        c.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill      = PatternFill("solid", start_color=cfg["dark"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _bdr()
        c.number_format = "0.00"

    rt_letter = get_column_letter(row_total_col)
    first_letter = get_column_letter(first_sys_col)
    last_letter  = get_column_letter(last_sys_col)
    c = ws.cell(1, row_total_col, f"=SUM({first_letter}1:{last_letter}1)")
    c.font = Font(name="Arial", bold=True, size=11, color="1A5C9E")
    c.fill = PatternFill("solid", start_color="EEF4FF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _bdr()
    c.number_format = "0.00"

    c = ws.cell(1, gbp_col, f"=SUM({rt_letter}1*10)")
    c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _bdr()
    c.number_format = "0.00"

    # ── Row 2: Headers ────────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 18.0
    headers = [
        (2,"Date",NAVY),(3,"Time",NAVY),(4,"League",NAVY),(5,"Home",NAVY),
        (6,"Away",NAVY),(7,"Market",NAVY),(8,"6G xG","0B5E6B"),(9,"Rule","0B5E6B"),
        (10,"Odds","1A5C9E"),(11,"Hist ROI","1A5C9E"),
    ]
    for i, (sys_name, cfg) in enumerate(sys_map.items()):
        headers.append((first_sys_col + i, sys_name, cfg["dark"]))
    headers += [
        (row_total_col,"Row Total","1A5C9E"),
        (month_col,"Month","1A5C9E"),
        (cum_col,"Cumulative",NAVY),
        (gbp_col,"\u00a3",NAVY),
    ]
    for col_num, label, bg in headers:
        c = ws.cell(2, col_num, label)
        c.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _bdr()

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, sig in enumerate(signals):
        row = i + 3
        ws.row_dimensions[row].height = 16.5
        cfg     = sys_map.get(sig.system)
        if cfg is None:
            continue
        light   = cfg["light"]; dark = cfg["dark"]; fg = cfg["fg"]; sys_col = cfg["col"]

        jk_bg = JK_ODD_BG if i % 2 == 0 else JK_EVEN_BG

        date_v = _fmt_date(sig.date)
        time_s = _fmt_time(sig.time)
        roi_s  = f"+{sig.hist_roi:.2f}%" if sig.hist_roi >= 0 else f"{sig.hist_roi:.2f}%"
        rule_s = sig.rule.replace(" | QUALIFIES","").replace(" | BUFFER"," \u26a0")

        # B-F: Calibri, size 11, BLACK font — plain data columns, no coloured fill
        dc = _plain(ws, row, 2, date_v, 11, "left")
        dc.number_format = "mm-dd-yy" if isinstance(date_v, datetime) else "General"
        _plain(ws, row, 3, time_s,     11, "center")
        _plain(ws, row, 4, sig.league, 11, "left")
        _plain(ws, row, 5, sig.home,   11, "left")
        _plain(ws, row, 6, sig.away,   11, "left")

        # G: dark bg, white, bold=True, size 11
        _c(ws, row, 7, sig.system, dark, "FFFFFF", True, 11, "center")
        # H: system light bg, system fg, bold=True, number_format 0.00
        _c(ws, row, 8, sig.xg_value, light, fg, True, 11, "center", number_format="0.00")
        # I: system light bg, system fg, bold=False, size 11
        _c(ws, row, 9, rule_s, light, fg, False, 11, "center")

        # J: Odds — alternating bg, 1A5C9E fg, bold=True, number_format 0.00
        is_btd    = sig.system == "Back the Draw"
        is_buffer = is_btd and sig.odds < 3.60
        _c(ws, row, 10, sig.odds,
           "FFF0DC" if is_buffer else jk_bg,
           "B35C00" if is_buffer else J_FG,
           True, 11, "center", number_format="0.00")

        # K: Hist ROI — alternating bg, 0B5E6B fg, bold=True, size 11
        _c(ws, row, 11, roi_s, jk_bg, K_FG_POS, True, 11, "center")

        # L-O (or L-P for single-system Test sheet): system result columns, number_format 0.00
        for col_num in range(first_sys_col, last_sys_col + 1):
            c = ws.cell(row, col_num)
            if col_num == sys_col:
                c.fill = PatternFill("solid", start_color=light)
                c.font = Font(name=BODY_FONT, size=11, color=ACTIVE_FG)
            else:
                c.fill = PatternFill("solid", start_color=INACT_BG)
                c.font = Font(name=BODY_FONT, size=11, color=INACT_FG)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = _bdr()
            c.number_format = "0.00"

        # Row Total
        c = ws.cell(row, row_total_col, f"=SUM({first_letter}{row}:{last_letter}{row})")
        c.font = Font(name="Arial", bold=True, size=11, color="1A5C9E")
        c.fill = PatternFill("solid", start_color="EEF4FF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _bdr()
        c.number_format = "0.00"

        # Month
        month_letter = get_column_letter(month_col)
        rt_ref = f"{rt_letter}{row}"
        c = ws.cell(row, month_col,
            f'=IF({rt_ref}=0,"",{rt_ref})' if row == 3 else f"=SUM({month_letter}{row-1}+{rt_ref})")
        c.font = Font(name="Arial", bold=True, size=11, color="0D2B55")
        c.fill = PatternFill("solid", start_color="F2F6FB")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _bdr()
        c.number_format = "0.00"

        # Cumulative
        cum_letter = get_column_letter(cum_col)
        c = ws.cell(row, cum_col,
            f'=IF({rt_ref}=0,"",{rt_ref})' if row == 3 else f"=SUM({cum_letter}{row-1}+{rt_ref})")
        c.font = Font(name="Arial", bold=True, size=11, color="0D2B55")
        c.fill = PatternFill("solid", start_color="F2F6FB")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _bdr()
        c.number_format = "0.00"

        # £
        c = ws.cell(row, gbp_col, f"=SUM({cum_letter}{row}*10)")
        c.font = Font(name="Arial", bold=True, size=11, color="0D2B55")
        c.fill = PatternFill("solid", start_color="F2F6FB")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _bdr()
        c.number_format = "0.00"

    ws.freeze_panes = "B3"


def export_to_excel(signals: List[BetSignal], filepath: str, date_str: str):
    wb = Workbook()

    live_signals = [s for s in signals if s.system in LIVE_SYS]
    test_signals = [s for s in signals if s.system in TEST_SYS]

    # ── Sheet 1: Selections (4 LIVE systems only) ─────────────────────────────
    ws1 = wb.active
    ws1.title = "Selections"
    _write_sheet(ws1, live_signals, date_str, LIVE_SYS, "FTS xG DAILY SELECTIONS")

    # ── Sheet 2: Test (Back the Draw only) ────────────────────────────────────
    ws_test = wb.create_sheet("Test")
    _write_sheet(ws_test, test_signals, date_str, TEST_SYS, "FTS xG TEST SELECTIONS (Back the Draw)")

    # ── Sheet 3: Results summary ──────────────────────────────────────────────
    from collections import Counter
    ws3 = wb.create_sheet("Results")
    ws3.sheet_view.showGridLines = False
    for col, w in {'A':22,'B':10,'C':10}.items():
        ws3.column_dimensions[col].width = w

    ws3.merge_cells("A1:C1")
    c = ws3['A1']
    c.value = "Daily Summary"
    c.font  = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill  = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 24

    mc = Counter(s.system for s in signals)
    all_sys = {**LIVE_SYS, **TEST_SYS}
    for i, sys_name in enumerate(["Lay U1.5","Back O2.5","Lay O3.5","FHG Lay U0.5","Back the Draw"]):
        r   = i + 2
        cfg = all_sys[sys_name]
        cnt = mc.get(sys_name, 0)
        status = " (TEST)" if sys_name == "Back the Draw" else ""
        for col, val, bg, fg, align in [
            (1, sys_name + status, cfg["dark"], "FFFFFF", "left"),
            (2, cnt,      cfg["light"], cfg["fg"], "center"),
            (3, "—",      cfg["light"], cfg["fg"], "center"),
        ]:
            c = ws3.cell(r, col, val)
            c.font      = Font(name="Arial", size=10, bold=(col==1), color=fg)
            c.fill      = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border    = _bdr()
        ws3.row_dimensions[r].height = 18

    tot = 7
    ws3.merge_cells(f"A{tot}:C{tot}")
    c = ws3[f'A{tot}']
    c.value = f"Total Selections: {len(signals)}  (Live: {len(live_signals)}  ·  Test: {len(test_signals)})"
    c.font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill  = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _bdr()

    wb.save(filepath)
